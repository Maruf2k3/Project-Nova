import io
import os
import csv
import uuid
from flask import json, render_template, request, redirect, url_for, flash, jsonify, Response
from flask_login import login_user, login_required, logout_user, current_user
from flask_jwt_extended import create_access_token, create_refresh_token, set_access_cookies, set_refresh_cookies, unset_jwt_cookies
from sqlalchemy import func
from models import User, MenuItem, Sale, Customer, Employee, Attendance,Group,GroupSale,GroupPayment,Inventory,InventoryLog,db
from werkzeug.utils import secure_filename
from datetime import datetime
import pytz
import openpyxl  # To handle Excel files


ALLOWED_EXTENSIONS = {'xls', 'xlsx'}

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

def register_routes(app):

    
    # 1. Error routes Handeled 
    @app.errorhandler(Exception)
    def handle_error(error):# Generic error handler for all error types
        # Default error code is 500 (Internal Server Error)
        error_code = getattr(error, 'code', 500)

        # Define error messages based on error code
        error_messages = {
            400: "Bad Request! The server could not understand your request.",
            404: "Oops! The page you're looking for does not exist.",
            500: "Sorry! Something went wrong on our end.",
            302: "Redirect! You are being redirected.",
        }

        # Get the error message or default to a generic one
        message = error_messages.get(error_code, "An unexpected error occurred.")

        return render_template('error.html', error_code=error_code, message=message), error_code

    # 2. Landing page route ("/")
    @app.route("/")
    def landing_page():
        if current_user.is_authenticated:
            return redirect(url_for('dashboard'))
        return render_template('base.html')
    
    # 3. Dashboard route
    @app.route("/dashboard")
    @login_required
    def dashboard():
        employees = Employee.query.all() if current_user.role == "admin" else []
        userName = current_user.username;
        return render_template("home.html", employees=employees , userName = userName)

    #---------------------------Auth--------------------#
    # 4.1 Login route with JWT tokens
    @app.route("/login", methods=['GET', 'POST'])
    def login():
        if current_user.is_authenticated:
            return redirect(url_for('dashboard'))

        if request.method == 'POST':
            username = request.form.get('username')
            password = request.form.get('password')
            user = User.query.filter_by(username=username).first()

            if user and user.check_password(password):
                login_user(user)
                # Generate JWT tokens
                access_token = create_access_token(identity=user.id)
                refresh_token = create_refresh_token(identity=user.id)

                response = redirect(url_for('dashboard'))
                set_access_cookies(response, access_token)
                set_refresh_cookies(response, refresh_token)
                return response
            else:
                flash('Invalid username or password.', 'error')

        return render_template('login.html')

    # 4.2 Register route
    # Updated Register Route with role restrictions
    @app.route("/register", methods=['GET', 'POST'])
    def register():
        if current_user.is_authenticated:
            return redirect(url_for('dashboard'))

        if request.method == 'POST':
            username = request.form.get('username')
            password = request.form.get('password')
            role = request.form.get('role')

            # Check if the username already exists
            if User.query.filter_by(username=username).first():
                flash('Username already exists.', 'error')
                return render_template('register.html')

            # Role restrictions logic
            if role == 'admin':
                # Check if an admin already exists
                if User.query.filter_by(role='admin').first():
                    flash('Only one admin is allowed.', 'error')
                    return render_template('register.html')

            elif role == 'manager':
                # Check if there are already two managers
                manager_count = User.query.filter_by(role='manager').count()
                if manager_count >= 2:
                    flash('Only two managers are allowed.', 'error')
                    return render_template('register.html')

            # Create a new user with the valid role
            new_user = User(username=username, role=role)
            new_user.set_password(password)
            db.session.add(new_user)
            db.session.commit()

            flash('Registered successfully. Please log in.', 'success')
            return redirect(url_for('login'))

        return render_template('register.html')

    # 4.3 Logout route with token invalidation
    @app.route("/logout")
    @login_required
    def logout():
        logout_user()
        response = redirect(url_for('login'))
        unset_jwt_cookies(response)  # Unset JWT tokens
        flash('Logged out successfully.', 'success')
        return response
    
    # 4.4 User Management route
    @app.route("/user-management", methods=['GET', 'POST'])
    @login_required
    def user_management():
        if current_user.role != 'admin':
            return redirect(url_for('dashboard'))

        if request.method == 'POST':
            action = request.form.get('action')
            user_id = request.form.get('user_id')

            if action == 'edit':
                # Editing an existing user
                user = User.query.get(user_id)

                if not user:
                    flash('User not found!', 'error')
                    return redirect(url_for('user_management'))

                # Update role
                user.role = request.form.get('role')

                # Check if password is provided and update if necessary
                password = request.form.get('password')
                if password:
                    user.set_password(password)  # Set new password if provided

                db.session.commit()
                flash('User updated successfully!', 'success')

            elif action == 'delete':
                # Deleting a user
                user = User.query.get(user_id)

                if not user:
                    flash('User not found!', 'error')
                    return redirect(url_for('user_management'))

                db.session.delete(user)
                db.session.commit()
                flash('User deleted successfully!', 'success')

        # Fetch all users for display
        users = User.query.all()
        return render_template('user_management.html', users=users)
    
    #---------------------------Menu----------------------#
    # 5.1 Menu Management (Upload Excel and CRUD)
    @app.route("/menu-management", methods=['GET', 'POST'])
    @login_required
    def menu_management():
        if current_user.role not in ['admin', 'manager']:
            return redirect(url_for('dashboard'))

        if request.method == 'POST':
            file = request.files.get('file')
            if file and allowed_file(file.filename):
                filename = secure_filename(file.filename)
                filepath = os.path.join('uploads', filename)
                file.save(filepath)

                # Read Excel and insert items into the database
                try:
                    workbook = openpyxl.load_workbook(filepath)
                    sheet = workbook.active
                    for row in sheet.iter_rows(min_row=2, values_only=True):  # Skip the header
                        name, price, category = row
                        # Check if the item already exists in the database to prevent duplicates
                        existing_item = MenuItem.query.filter_by(name=name).first()
                        if existing_item:
                            flash(f"Item '{name}' already exists in the database.", 'warning')
                        else:
                            new_item = MenuItem(name=name, price=float(price), category=category)
                            db.session.add(new_item)
                    db.session.commit()
                    flash('Menu uploaded and saved successfully!', 'success')
                except Exception as e:
                    flash(f"An error occurred while processing the file: {str(e)}", 'error')

                # Delete the file after successful processing
                if os.path.exists(filepath):
                    os.remove(filepath)
                    flash('Uploaded file deleted successfully.', 'info')

                return redirect(url_for('menu_management'))

        menu_items = MenuItem.query.all()
        return render_template('menu_management.html', menu_items=menu_items)

    # 5.2 Route for editing a menu item
    @app.route("/menu-management/edit/<int:item_id>", methods=['GET', 'POST'])
    @login_required
    def edit_menu_item(item_id):
        if current_user.role not in ['admin', 'manager']:
            return redirect(url_for('dashboard'))

        # Fetch the menu item by its ID
        menu_item = MenuItem.query.get_or_404(item_id)

        if request.method == 'POST':
            name = request.form.get('name')
            price = request.form.get('price')
            category = request.form.get('category')

            # Update the menu item with new values
            menu_item.name = name
            menu_item.price = float(price)
            menu_item.category = category

            db.session.commit()
            flash('Menu item updated successfully.', 'success')
            return redirect(url_for('menu_management'))

        return render_template('edit_menu_item.html', item=menu_item)

    # 5.3 Deleting Menu item
    @app.route("/menu-management/delete/<int:item_id>")
    @login_required
    def delete_menu_item(item_id):
        if current_user.role not in ['admin', 'manager']:
            return redirect(url_for('dashboard'))

        item = MenuItem.query.get_or_404(item_id)
        db.session.delete(item)
        db.session.commit()
        flash('Menu item deleted successfully!', 'success')
        return redirect(url_for('menu_management'))

    #-------------------------Bill-----------------------#
    # 6. Create Bill both either Customer or Group
    @app.route("/create-bill", methods=['GET', 'POST'])
    @login_required
    def create_bill():
        if current_user.role not in ['cashier', 'manager', 'admin']:
            return redirect(url_for('dashboard'))

        if request.method == 'POST':
            # Fetch form data
            customer_id = request.form.get('customer')
            group_id = request.form.get('group')  # Fetch group ID if selected
            note = request.form.get('note')
            discount = int(request.form.get('discount', 0))
            taxFromHtml = int(request.form.get('tax', 0))
            payment_method = request.form.get('payment-method', 'cash')

            # Handle cases when no customer or group is selected
            customer = None
            group = None

            if customer_id and customer_id != 'None':
                customer = Customer.query.get(customer_id)
            
            if group_id and group_id != 'None':
                group = Group.query.get(group_id)

            # Fetch items from form data (expecting JSON string of items from front-end)
            items = request.form.get('items')  # Expecting JSON string

            # Parse items as a list of dictionaries (item name, quantity, price)
            if items:
                try:
                    items = json.loads(items)
                except json.JSONDecodeError:
                    flash('Error parsing items.', 'error')
                    return redirect(url_for('create_bill'))

            # Ensure items contain 'price' and 'qty'
            if not all('price' in item and 'qty' in item for item in items):
                flash('Error: Items data is incomplete.', 'error')
                return redirect(url_for('create_bill'))

            # Calculate subtotal, tax, and grand total
            subtotal = sum(float(item['price']) * int(item['qty']) for item in items)
            tax = float(taxFromHtml / 100) * subtotal
            grand_total = subtotal + tax - discount

            # Generate a unique invoice number using uuid4
            invoice_number = str(uuid.uuid4())
            server = current_user.username
            current_time_uae = datetime.now(pytz.timezone("Asia/Dubai"))

            # Case 1: If a customer is selected
            if customer:
                new_sale = Sale(
                    invoice_number=invoice_number,
                    date=current_time_uae.strftime("%Y-%m-%d"),
                    time=current_time_uae.strftime("%H:%M:%S"),
                    items=json.dumps(items),  # Storing the items as JSON in the DB
                    subtotal=subtotal,
                    tax=tax,
                    grand_total=grand_total,
                    discount=discount,
                    notes=note,
                    server=server,  # Store server name
                    customer_id=customer.id,  # Store customer_id only if customer exists
                    payment_method=payment_method
                )
                # Add and commit the sale
                db.session.add(new_sale)

            # Case 2: If a group is selected
            elif group:
                new_group_sale = GroupSale(
                    invoice_number=invoice_number,
                    date=current_time_uae.strftime("%Y-%m-%d"),
                    time=current_time_uae.strftime("%H:%M:%S"),
                    items=json.dumps(items),
                    subtotal=subtotal,
                    tax=tax,
                    grand_total=grand_total,
                    discount=discount,
                    group_id=group.id,  # Store group_id
                    payment_method=payment_method,
                    server=server,
                    notes=note
                )
                # Update the group's total due
                group.total_due += grand_total
                db.session.add(new_group_sale)
            
             # Case 3: Walk-in (no customer or group selected)
            else:
                new_sale = Sale(
                    invoice_number=invoice_number,
                    date=current_time_uae.strftime("%Y-%m-%d"),
                    time=current_time_uae.strftime("%H:%M:%S"),
                    items=json.dumps(items),
                    subtotal=subtotal,
                    tax=tax,
                    grand_total=grand_total,
                    discount=discount,
                    notes=note,
                    server=server,  # Store server name
                    customer_id=None,  # No customer or group, walk-in
                    payment_method=payment_method
                )
                # Add and commit the sale for walk-in
                db.session.add(new_sale)

            # Commit the transaction and emit event for all connected clients
            db.session.commit()

            

            flash('Bill created successfully!', 'success')
            return redirect(url_for('dashboard'))

        # Load menu items, customers, and groups for the dropdowns
        customers = Customer.query.all() if Customer.query.count() > 0 else []
        groups = Group.query.all() if Group.query.count() > 0 else []
        menu_items = MenuItem.query.all()  # Ensure MenuItem is loaded

        return render_template('create_bill.html', customers=customers, groups=groups, menu_items=menu_items)

    #-------------------------------------Sales Reports----------------------------
    # 7.1 #To convert json to obj
    @app.template_filter('fromjson')
    def fromjson_filter(s):
        try:
            return json.loads(s)
        except (ValueError, TypeError):
            return {}
    # 7.2 Sales Report For Customer
    @app.route("/sales-reports", methods=['GET', 'POST'])
    @login_required
    def sales_reports():
        if current_user.role not in ['admin', 'manager']:
            return redirect(url_for('dashboard'))

        page = request.args.get('page', 1, type=int)
        per_page = 15
        sales = []
        start_date_str = None
        end_date_str = None

        if request.method == 'POST':
            # Get dates from form submission
            start_date_str = request.form.get('start_date')
            end_date_str = request.form.get('end_date')
        else:
            # Get dates from query parameters (for pagination links)
            start_date_str = request.args.get('start_date')
            end_date_str = request.args.get('end_date')

        if start_date_str and end_date_str:
            # Validate the dates
            try:
                datetime.strptime(start_date_str, '%Y-%m-%d')
                datetime.strptime(end_date_str, '%Y-%m-%d')
            except ValueError:
                flash('Invalid date format.', 'error')
                return redirect(url_for('sales_reports'))

            # Fetch sales data within the date range with pagination
            sales = Sale.query.filter(
                func.date(Sale.date) >= start_date_str.strip(),
                func.date(Sale.date) <= end_date_str.strip()
            ).order_by(Sale.date.desc()).paginate(page=page, per_page=per_page)

            if not sales.items and page == 1:
                flash('No sales found within the selected date range.', 'error')
        else:
            # If no dates provided, render the empty form
            sales = None

        return render_template('sales_reports.html', sales=sales, start_date=start_date_str, end_date=end_date_str)

    
    # 7.3 Predefined Sales Report Route for different date ranges
    @app.route("/sales-reports/predefined/<string:range_type>", methods=['GET'])
    @login_required
    def sales_reports_predefined(range_type):
        if current_user.role not in ['admin', 'manager']:
            return redirect(url_for('dashboard'))

        page = request.args.get('page', 1, type=int)
        per_page = 15

        # Determine date range based on predefined range_type
        today = datetime.now().date()
        if range_type == 'today':
            start_date = end_date = today
        elif range_type == 'month':
            start_date = today.replace(day=1)
            end_date = today
        elif range_type == 'year':
            start_date = today.replace(month=1, day=1)
            end_date = today
        else:
            flash('Invalid range type.', 'error')
            return redirect(url_for('sales_reports'))

        # Convert dates to string format for querying
        start_date_str = start_date.strftime('%Y-%m-%d')
        end_date_str = end_date.strftime('%Y-%m-%d')

        # Query the sales data with pagination
        sales = Sale.query.filter(
            func.date(Sale.date) >= start_date_str,
            func.date(Sale.date) <= end_date_str
        ).order_by(Sale.date.desc()).paginate(page=page, per_page=per_page)

        if not sales.items and page == 1:
            flash('No sales found for the selected range.', 'error')

        # Render template with pagination and sales data
        return render_template('sales_reports.html', sales=sales, range_type=range_type , start_date=start_date, end_date=end_date)


    # 7.4 Route to Download Sales Report Customer
    # Updated Download Sales Report Route
    @app.route("/download-sales-report", methods=['GET'])
    @login_required
    def download_sales_report():
        if current_user.role not in ['admin', 'manager']:
            return redirect(url_for('dashboard'))

        start_date_str = request.args.get('start_date')
        end_date_str = request.args.get('end_date')

        if not start_date_str or not end_date_str:
            flash('Invalid date range for report download.', 'error')
            return redirect(url_for('sales_reports'))

        try:
            # Ensure dates are in the correct format
            if start_date_str:
                datetime.strptime(start_date_str, '%Y-%m-%d')
            if end_date_str:
                datetime.strptime(end_date_str, '%Y-%m-%d')
        except ValueError:
            flash('Invalid date format.', 'error')
            return redirect(url_for('sales_reports'))

        # Fetch sales data within the date range
        sales = Sale.query.filter(
            func.date(Sale.date) >= start_date_str.strip(),
            func.date(Sale.date) <= end_date_str.strip()
        ).all()

        if not sales:
            flash('No sales found within the selected date range.', 'error')
            return redirect(url_for('sales_reports'))

        # Generate the CSV data in memory
        output = io.StringIO()
        csv_writer = csv.writer(output)
        csv_writer.writerow(['Invoice Number', 'Date', 'Time', 'Customer', 'Items', 'Subtotal', 'Tax', 'Discount', 'Grand Total', 'Served By', 'Payment Method', 'Notes'])

        for sale in sales:
            customer_name = sale.customer.name if sale.customer else "No Customer"
            items = json.loads(sale.items)

            item_list = "; ".join([f"{item['name']} (Qty: {item.get('qty', 1)})" for item in items])

            csv_writer.writerow([
                sale.invoice_number,
                sale.date,
                sale.time,
                customer_name,
                item_list,
                sale.subtotal,
                sale.tax,
                sale.discount,
                sale.grand_total,
                sale.server,
                sale.payment_method,
                sale.notes
            ])

        # Move to the beginning of the StringIO object
        output.seek(0)

        # Create a Response object with the CSV data
        return Response(
            output.getvalue(),
            mimetype='text/csv',
            headers={
                'Content-Disposition': f'attachment;filename=salesReport_{start_date_str}_{end_date_str}.csv'
            }
        )



    #-----------------------CRM-------------------#
    # 8.1 Customer Management route
    @app.route("/customer-management", methods=['GET', 'POST'])
    @login_required
    def customer_management():
        if current_user.role not in ['admin', 'manager']:
            return redirect(url_for('dashboard'))

        if request.method == 'POST':
            action = request.form.get('action')

            if action == 'add':
                # Adding a new customer
                name = request.form.get('name')
                phone_number = request.form.get('phone_number')
                new_customer = Customer(name=name, phone_number=phone_number)
                db.session.add(new_customer)
                db.session.commit()
                flash('Customer added successfully!', 'success')
            elif action == 'edit':
                # Editing an existing customer
                customer_id = request.form.get('customer_id')
                customer = Customer.query.get(customer_id)
                customer.name = request.form.get('name')
                customer.phone_number = request.form.get('phone_number')
                db.session.commit()
                flash('Customer updated successfully!', 'success')
            elif action == 'delete':
                # Deleting a customer
                customer_id = request.form.get('customer_id')
                customer = Customer.query.get(customer_id)
                db.session.delete(customer)
                db.session.commit()
                flash('Customer deleted successfully!', 'success')

        customers = Customer.query.all()
        return render_template('customer_management.html', customers=customers)


    # 8.2 Route for viewing customer history
    @app.route("/customer-history", methods=['GET', 'POST'])
    @login_required
    def customer_history():
        if current_user.role not in ['admin', 'manager']:
            return redirect(url_for('dashboard'))

        customers = Customer.query.all()
        sales_history = None
        if request.method == 'POST':
            customer_id = request.form.get('customer_id')
            history_type = request.form.get('history_type')
            
            customer = Customer.query.get(customer_id)
            if history_type == 'date_range':
                start_date = request.form.get('start_date')
                end_date = request.form.get('end_date')
                sales_history = Sale.query.filter(Sale.customer_id == customer_id, Sale.date.between(start_date, end_date)).all()
            else:
                sales_history = Sale.query.filter_by(customer_id=customer_id).all()

        return render_template('customer_history.html', customers=customers, sales_history=sales_history)
    
    # 8.3 sales Report Customer | Used for the downlaod each Bill via JS
    @app.route("/get-sale/<int:sale_id>", methods=['GET'])
    @login_required
    def get_sale(sale_id):
        sale = Sale.query.get_or_404(sale_id)
        items = json.loads(sale.items)

        # Prepare sale data as JSON
        sale_data = {
            "invoice_number": sale.invoice_number,
            "date": sale.date,
            "time": sale.time,
            "items": items,
            "subtotal": sale.subtotal,
            "tax": sale.tax,
            "discount": sale.discount,
            "grand_total": sale.grand_total,
            "server": sale.server,
            "payment_method": sale.payment_method,
            "notes": sale.notes,
            "customer": sale.customer.name if sale.customer else 'No Customer'
        }

        return jsonify(sale_data)
    
    # 8.4 Route to delete a customer sale
    @app.route('/delete-sale/<int:sale_id>', methods=['POST'])
    @login_required
    def delete_sale(sale_id):
        if current_user.role not in ['admin']:
            return redirect(url_for('dashboard'))
        sale = Sale.query.get_or_404(sale_id)
        db.session.delete(sale)
        db.session.commit()
        flash('Sale deleted successfully!', 'success')
        return redirect(url_for('sales_reports'))

    #--------------------------Employee Managment--------------------#
    # 9.1 Employee Dashboard to display all employees
    @app.route("/employees", methods=["GET", "POST"])
    @login_required
    def employee_dashboard():
        employees = Employee.query.all()

        # Adding clock-in status to employee object
        for employee in employees:
            employee.clocked_in = Attendance.query.filter_by(employee_id=employee.id, clock_out=None).first() is not None

        if request.method == 'POST':
            # Handle adding new employees
            name = request.form.get('name')
            position = request.form.get('position')
            contact_info = request.form.get('contact_info')

            new_employee = Employee(name=name, position=position, contact_info=contact_info)
            db.session.add(new_employee)
            db.session.commit()
            flash('Employee added successfully!', 'success')
            return redirect(url_for('employee_dashboard'))

        return render_template('employee_dashboard.html', employees=employees)

    # 9.2 Add or Edit Employee (Form View)
    @app.route("/employee/add", methods=["GET", "POST"])
    @app.route("/employee/edit/<int:employee_id>", methods=["GET", "POST"])
    @login_required
    def add_edit_employee(employee_id=None):
        employee = Employee.query.get(employee_id) if employee_id else None

        if request.method == "POST":
            name = request.form.get('name')
            position = request.form.get('position')
            contact_info = request.form.get('contact_info')

            if employee:
                # Update existing employee
                employee.name = name
                employee.position = position
                employee.contact_info = contact_info
                flash('Employee updated successfully!', 'success')
            else:
                # Add new employee
                new_employee = Employee(name=name, position=position, contact_info=contact_info)
                db.session.add(new_employee)
                flash('Employee added successfully!', 'success')

            db.session.commit()
            return redirect(url_for('employee_dashboard'))

        return render_template('add_edit_employee.html', employee=employee)

    # 9.3 Delete Employee
    @app.route("/employee/delete/<int:employee_id>", methods=["POST"])
    @login_required
    def delete_employee(employee_id):
        employee = Employee.query.get_or_404(employee_id)
        db.session.delete(employee)
        db.session.commit()
        flash('Employee deleted successfully!', 'success')
        return redirect(url_for('employee_dashboard'))
    
    #9.4 Route for the caheir Only Dashboard
    @app.route("/e", methods=["GET"])
    @login_required
    def e():
        
        em = Employee.query.all()

        # Adding clock-in status to employee object
        for employee in em:
            employee.clocked_in = Attendance.query.filter_by(employee_id=employee.id, clock_out=None).first() is not None

        return render_template('casheirDashboard.html', employees=em)

    # 9.5 Clock-in route for employee Attendence
    @app.route("/employee/clock-in/<int:employee_id>", methods=["POST"])
    @login_required
    def clock_in(employee_id):
        employee = Employee.query.get_or_404(employee_id)

        clock_in_time = datetime.now(pytz.timezone("Asia/Dubai"))
        attendance = Attendance(employee_id=employee.id, clock_in=clock_in_time)

        db.session.add(attendance)
        db.session.commit()

       


        flash(f'{employee.name} clocked in successfully!', 'success')
        return redirect(request.referrer or url_for('dashboard'))

    # 9.6 Clock-out route For Employee Attendence
    @app.route("/employee/clock-out/<int:employee_id>", methods=["POST"])
    @login_required
    def clock_out(employee_id):
        attendance = Attendance.query.filter_by(employee_id=employee_id, clock_out=None).first()
        if attendance:
            clock_out_time = datetime.now(pytz.timezone("Asia/Dubai"))
            attendance.clock_out = clock_out_time
            db.session.commit()
            

        else:
            flash('No clock-in record found for the employee.', 'error')
        return redirect(request.referrer or url_for('dashboard'))

    # 9.7 Employee Profile View
    @app.route("/employee/profile/<int:employee_id>", methods=["GET"])
    @login_required
    def employee_profile(employee_id):
        employee = Employee.query.get_or_404(employee_id)
        attendance_records = Attendance.query.filter_by(employee_id=employee_id).all()
        return render_template('employee_profile.html', employee=employee, attendance_records=attendance_records)

    # 9.8 Attendance Selection Page (select employee to view attendance)
    @app.route("/attendance/select", methods=["GET", "POST"])
    @login_required
    def attendance_selection():
        employees = Employee.query.all()

        if request.method == "POST":
            employee_id = request.form.get("employee_id")
            return redirect(url_for('employee_attendance', employee_id=employee_id))

        return render_template("attendance_selection.html", employees=employees)

    # 9.9 View Attendance Records (for selected employee)
    @app.route("/employee/attendance/<int:employee_id>", methods=["GET"])
    @login_required
    def employee_attendance(employee_id):
        employee = Employee.query.get_or_404(employee_id)
        attendance_records = Attendance.query.filter_by(employee_id=employee_id).all()
        return render_template('employee_attendance.html', employee=employee, attendance_records=attendance_records)

    #---------------------Group Managment--------------------#
    #11.1 To get Group sales Report
    @app.route("/group-sales-reports", methods=['GET', 'POST'])
    @login_required
    def group_sales_reports():
        if current_user.role not in ['admin', 'manager']:
            return redirect(url_for('dashboard'))

        group_sales = []
        start_date_str = None
        end_date_str = None
        group_id = None

        if request.method == 'POST':
            start_date_str = request.form.get('start_date')
            end_date_str = request.form.get('end_date')
            group_id = request.form.get('group_id')

            try:
                # Ensure dates are in the correct format
                datetime.strptime(start_date_str, '%Y-%m-%d')
                datetime.strptime(end_date_str, '%Y-%m-%d')
            except ValueError:
                flash('Invalid date format.', 'error')
                return redirect(url_for('group_sales_reports'))

            # Fetch group sales data within the date range
            query = GroupSale.query.filter(
                func.date(GroupSale.date) >= start_date_str.strip(),
                func.date(GroupSale.date) <= end_date_str.strip()
            )
            if group_id:
                query = query.filter(GroupSale.group_id == group_id)

            group_sales = query.all()

            if not group_sales:
                flash('No group sales found within the selected date range.', 'error')
                return render_template('group_sales_reports.html', group_sales=group_sales)

        groups = Group.query.all()  # Load all groups for selection
        return render_template('group_sales_reports.html', group_sales=group_sales, groups=groups, start_date=start_date_str, end_date=end_date_str, group_id=group_id)

    # 11.1.1 Predefined Group Sales Report Route for different date ranges
    @app.route("/group-sales-reports/predefined/<string:range_type>", methods=['GET'])
    @login_required
    def group_sales_reports_predefined(range_type):
        if current_user.role not in ['admin', 'manager']:
            return redirect(url_for('dashboard'))

        today = datetime.now().date()
        if range_type == 'today':
            start_date = end_date = today
        elif range_type == 'month':
            start_date = today.replace(day=1)
            end_date = today
        elif range_type == 'year':
            start_date = today.replace(month=1, day=1)
            end_date = today
        else:
            flash('Invalid range type.', 'error')
            return redirect(url_for('group_sales_reports'))

        start_date_str = start_date.strftime('%Y-%m-%d')
        end_date_str = end_date.strftime('%Y-%m-%d')

        group_sales = GroupSale.query.filter(
            func.date(GroupSale.date) >= start_date_str,
            func.date(GroupSale.date) <= end_date_str
        ).all()

        groups = Group.query.all()
        return render_template('group_sales_reports.html', group_sales=group_sales, groups=groups, start_date=start_date_str, end_date=end_date_str, range_type=range_type)
    
    # 11.2 Root route for groups
    @app.route('/group-management', methods=['GET'])
    @login_required
    def group_management():
        if current_user.role not in ['admin', 'manager']:
            return redirect(url_for('dashboard'))
        userRole = current_user.role

        groups = Group.query.all()  # Fetch all groups from the database
        return render_template('group_management.html', groups=groups , role = userRole)
    
    # 11.3 For creating new groups
    @app.route('/groups/create', methods=['POST'])
    @login_required
    def create_group():
        if current_user.role not in ['admin', 'manager']:
            return redirect(url_for('dashboard'))

        name = request.form.get('name')
        if not name:
            flash('Group name is required!', 'error')
            return redirect(url_for('group_management'))

        # Create and save the new group
        new_group = Group(name=name)
        db.session.add(new_group)
        db.session.commit()

        flash(f'Group "{name}" created successfully!', 'success')
        return redirect(url_for('group_management'))
    
    # 11.4 edit Groups
    @app.route('/groups/edit/<int:group_id>', methods=['GET', 'POST'])
    @login_required
    def edit_group(group_id):
        group = Group.query.get_or_404(group_id)

        if request.method == 'POST':
            group.name = request.form.get('name')
            db.session.commit()
            flash(f'Group "{group.name}" updated successfully.', 'success')
            return redirect(url_for('group_management'))

        return render_template('edit_group.html', group=group)

    # 11.5 Delete Group
    @app.route('/groups/delete/<int:group_id>', methods=['POST'])
    @login_required
    def delete_group(group_id):
        group = Group.query.get_or_404(group_id)

        # Delete the group
        db.session.delete(group)
        db.session.commit()

        flash(f'Group "{group.name}" deleted successfully.', 'success')
        return redirect(url_for('group_management'))
    
    # 11.6 Group payment Updates
    @app.route('/groups/payment/<int:group_id>', methods=['POST'])
    @login_required
    def settle_group_payment(group_id):
        group = Group.query.get_or_404(group_id)
        payment = float(request.form.get('payment'))

        # Update group's payment details
        group.total_paid += payment
        group.total_due -= payment
        current_time_uae = datetime.now(pytz.timezone("Asia/Dubai"))
        

        # Create a payment record
        new_payment = GroupPayment(group_id=group.id, amount_paid=payment , date_paid = current_time_uae)
        db.session.add(new_payment)

        db.session.commit()

        flash(f"Payment of {payment} settled for group {group.name}.", 'success')
        return redirect(url_for('group_management'))
    
    # 11.7 Payment history View for Group
    @app.route("/group/payment-history/<int:group_id>", methods=['GET'])
    @login_required
    def group_payment_history(group_id):
        group = Group.query.get_or_404(group_id)
        payments = GroupPayment.query.filter_by(group_id=group_id).all()
        return render_template('group_payment_history.html', group=group, payments=payments)
    
    # 11.8 Route to Download Group Sales Report
    @app.route("/download-group-sales-report", methods=['GET'])
    @login_required
    def download_group_sales_report():
        if current_user.role not in ['admin', 'manager']:
            return redirect(url_for('dashboard'))

        start_date_str = request.args.get('start_date')
        end_date_str = request.args.get('end_date')
        group_id = request.args.get('group_id')

        try:
            # Ensure dates are in the correct format
            datetime.strptime(start_date_str, '%Y-%m-%d')
            datetime.strptime(end_date_str, '%Y-%m-%d')
        except ValueError:
            flash('Invalid date format.', 'error')
            return redirect(url_for('group_sales_reports'))

        # Fetch group sales data within the date range
        query = GroupSale.query.filter(
            func.date(GroupSale.date) >= start_date_str.strip(),
            func.date(GroupSale.date) <= end_date_str.strip()
        )
        if group_id:
            query = query.filter(GroupSale.group_id == group_id)

        group_sales = query.all()

        if not group_sales:
            flash('No group sales found to generate the report.', 'error')
            return redirect(url_for('group_sales_reports'))

        # Create CSV file in memory
        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerow(['Invoice Number', 'Date', 'Time', 'Group', 'Items', 'Subtotal', 'Tax', 'Discount', 'Grand Total', 'Served By', 'Payment Method', 'Notes'])

        for sale in group_sales:
            items = json.loads(sale.items)
            items_str = ', '.join([f"{item['name']} (Qty: {item.get('qty', 1)})" for item in items])
            writer.writerow([sale.invoice_number,
                sale.date,
                sale.time,
                sale.group.name,
                items_str,
                sale.subtotal,
                sale.tax,
                sale.discount,
                sale.grand_total,
                sale.server,
                sale.payment_method,
                sale.notes])

        output.seek(0)

        return Response(output, mimetype='text/csv', headers={
            'Content-Disposition': f'attachment;filename=group_sales_report_{start_date_str}_to_{end_date_str}.csv'
        })
    
    # 11.9 Download payment history of the Group
    @app.route("/group/payment-history/download/<int:group_id>", methods=['GET'])
    @login_required
    def download_group_payment_history(group_id):
        group = Group.query.get_or_404(group_id)
        payments = GroupPayment.query.filter_by(group_id=group_id).all()

        # Generate CSV data in memory
        output = io.StringIO()
        csv_writer = csv.writer(output)
        # Add a row for the group name
        csv_writer.writerow([f'Payment History for {group.name}'])
        csv_writer.writerow([])  # Empty row for spacing
        csv_writer.writerow(['Payment Amount', 'Date Paid'])

        for payment in payments:
            csv_writer.writerow([payment.amount_paid, payment.date_paid.strftime('%Y-%m-%d %H:%M:%S')])

        output.seek(0)

        # Create a Response object to serve the CSV file
        return Response(
            output.getvalue(),
            mimetype="text/csv",
            headers={"Content-Disposition": f"attachment;filename={group.name}_payment_history.csv"}
        )

    # 11.10 Reset the paid amount to zero
    @app.route('/groups/reset-paid/<int:group_id>', methods=['POST'])
    @login_required
    def reset_total_paid(group_id):
        if current_user.role not in ['admin']:
            return redirect(url_for('dashboard'))
        group = Group.query.get_or_404(group_id)

        # Reset the total paid amount for the group
        group.total_paid = 0.0
        group.total_due = 0.00
        db.session.commit()
        

        flash(f"Total paid amount for group {group.name} has been reset to zero.", 'success')
        return redirect(url_for('group_management'))

    # 11.11 Route to delete a group sale
    @app.route('/delete-group-sale/<int:group_sale_id>', methods=['POST'])
    @login_required
    def delete_group_sale(group_sale_id):
        if current_user.role not in ['admin']:
            return redirect(url_for('dashboard'))
        group_sale = GroupSale.query.get_or_404(group_sale_id)
        db.session.delete(group_sale)
        db.session.commit()
        flash('Group sale deleted successfully!', 'success')
        return redirect(url_for('group_sales_reports'))

    # 11.12 Get Sales fo the group
    @app.route("/get-group-sale/<int:group_sale_id>", methods=['GET'])
    @login_required
    def get_group_sale(group_sale_id):
        group_sale = GroupSale.query.get_or_404(group_sale_id)
        items = json.loads(group_sale.items)

        # Prepare sale data as JSON
        sale_data = {
            "invoice_number": group_sale.invoice_number,
            "date": group_sale.date,
            "time": group_sale.time,
            "items": items,
            "subtotal": group_sale.subtotal,
            "tax": group_sale.tax,
            "discount": group_sale.discount,
            "grand_total": group_sale.grand_total,
            "server": group_sale.server,
            "payment_method": group_sale.payment_method,
            "notes": group_sale.notes
        }

        return jsonify(sale_data)

    #12.1 Add Inventory Via text | Single Items at a time
    @app.route('/inventory/add', methods=['POST'])
    @login_required
    def add_inventory():
        if current_user.role not in ['admin']:
            return redirect(url_for('inventory_page'))

        # Extract form data
        name = request.form.get('name').strip()
        quantity = request.form.get('quantity')
        unit = request.form.get('unit')

        # Validate quantity
        try:
            quantity = float(quantity)
            if quantity <= 0:
                flash('Quantity must be greater than zero.', 'error')
                return redirect(url_for('inventory_page'))
        except ValueError:
            flash('Invalid quantity value.', 'error')
            return redirect(url_for('inventory_page'))

        # Check if the item already exists
        existing_item = Inventory.query.filter_by(name=name).first()
        if existing_item:
            flash(f"Item '{name}' already exists in the inventory. Consider editing the item instead.", 'error')
        else:
            # Add new item to inventory
            new_item = Inventory(name=name, quantity=quantity, unit=unit)
            db.session.add(new_item)
            db.session.commit()
            flash(f"Added new item '{name}' to the inventory.", 'success')

        return redirect(url_for('inventory_page'))

    #12.2 Add the Logs for what have been used by date Multiple at a time
    @app.route('/inventory/use', methods=['POST'])
    @login_required
    def use_inventory():
        item_ids = request.form.getlist('item_id[]')  # List of item IDs
        used_quantities = request.form.getlist('used_quantity[]')  # List of quantities
        usage_dates = request.form.getlist('usage_date[]')  # List of dates

        # Loop through each item
        for i in range(len(item_ids)):
            item_id = item_ids[i]
            used_quantity = float(used_quantities[i])
            usage_date = usage_dates[i]

            # Parse the date
            try:
                usage_date = datetime.strptime(usage_date, '%Y-%m-%d')
            except ValueError:
                flash('Invalid date format. Please use YYYY-MM-DD.', 'error')
                return redirect(url_for('inventory_page'))

            # Fetch the item from the database
            item = Inventory.query.get(item_id)
            if item:
                # Check if enough quantity is available
                if item.quantity >= used_quantity:
                    item.subtract_quantity(used_quantity)  # Update the quantity
                    # Log the usage
                    log = InventoryLog(item_id=item.id, used_quantity=used_quantity, date=usage_date)
                    db.session.add(log)
                    db.session.commit()
                    flash(f"Used {used_quantity} {item.unit} of {item.name} on {usage_date.strftime('%Y-%m-%d')}.", 'success')
                else:
                    flash(f"Cannot subtract {used_quantity}. Only {item.quantity} {item.unit} available for {item.name}.", 'error')
            else:
                flash(f"Item with ID {item_id} not found.", 'error')

        return redirect(url_for('inventory_logs'))

    #12.3 Add inventory bulk via excel
    @app.route('/inventory/upload', methods=['POST'])
    @login_required
    def upload_inventory():
        if current_user.role not in ['admin']:
            
            return redirect(url_for('inventory_page'))

        file = request.files.get('file')
        
        if file and allowed_file(file.filename):
            filename = secure_filename(file.filename)
            filepath = os.path.join('uploads', filename)
            file.save(filepath)

            # Load Excel file
            try:
                workbook = openpyxl.load_workbook(filepath)
                sheet = workbook.active
                for row in sheet.iter_rows(min_row=2, values_only=True):  # Skip header
                    name, quantity, unit = row
                    
                    # Check if the item already exists in the inventory
                    existing_item = Inventory.query.filter_by(name=name).first()
                    
                    if existing_item:
                        # If the item exists, update the quantity
                        existing_item.add_quantity(float(quantity))
                    else:
                        # If not, add a new inventory item
                        new_item = Inventory(name=name, quantity=float(quantity), unit=unit)
                        db.session.add(new_item)
                
                db.session.commit()
                flash('Inventory updated successfully from Excel file.', 'success')
            except Exception as e:
                flash(f"Error processing the file: {str(e)}", 'error')

            # Clean up the file after processing
            if os.path.exists(filepath):
                os.remove(filepath)
                flash('Uploaded file deleted after processing.', 'info')

        return redirect(url_for('inventory_page'))

    @app.route('/inventory')
    @login_required
    def inventory_page():
        inventory_items = Inventory.query.all()
        return render_template('inventory.html', inventory_items=inventory_items)

    @app.route('/inventory/logs', defaults={'page': 1})
    @app.route('/inventory/logs/page/<int:page>')
    @login_required
    def inventory_logs(page):
        per_page = 10  # Number of logs to display per page
        logs = InventoryLog.query.order_by(InventoryLog.date.desc()).paginate(page=page, per_page=per_page, error_out=False)
        
        return render_template('inventory_logs.html', logs=logs)

    
    @app.route('/inventory/log_usage')
    @login_required
    def log_usage_page():
        # Retrieve all inventory items to populate the dropdown in the form
        inventory_items = Inventory.query.all()
        return render_template('log_usage.html', inventory_items=inventory_items)

    @app.route('/inventory/delete/<int:item_id>', methods=['POST'])
    @login_required
    def delete_inventory(item_id):
        if current_user.role not in ['admin']:
            return redirect(url_for('inventory_page'))

        item = Inventory.query.get(item_id)
        if item:
            db.session.delete(item)
            db.session.commit()
            flash(f"Deleted item '{item.name}' from the inventory.", 'success')
        else:
            flash("Item not found.", 'error')
        return redirect(url_for('inventory_page'))
    
    
    @app.route('/inventory/edit', methods=['POST'])
    @login_required
    def edit_inventory():
        if current_user.role not in ['admin']:
            return redirect(url_for('inventory_page'))

        item_id = request.form.get('item_id')
        name = request.form.get('name').strip()
        quantity = float(request.form.get('quantity'))
        unit = request.form.get('unit')

        item = Inventory.query.get(item_id)
        if item:
            item.name = name
            item.quantity = quantity
            item.unit = unit
            db.session.commit()
            flash(f"Updated item '{item.name}' in the inventory.", 'success')
        else:
            flash("Item not found.", 'error')

        return redirect(url_for('inventory_page'))



    # Route to download the current inventory as CSV
    @app.route('/download-inventory', methods=['GET'])
    @login_required
    def download_inventory():
        if current_user.role not in ['admin']:
            flash(f'Your Current Role is: {current_user.role}. You do not have permission to download the inventory.', 'error')
            return redirect(url_for('inventory_page'))

        # Fetch all inventory items from the database
        inventory_items = Inventory.query.all()

        if not inventory_items:
            flash('No inventory items found.', 'error')
            return redirect(url_for('inventory_page'))

        # Generate the CSV data in memory
        output = io.StringIO()
        csv_writer = csv.writer(output)

        # Write CSV header
        csv_writer.writerow(['Item Name', 'Quantity', 'Unit'])

        # Write inventory data rows
        for item in inventory_items:
            csv_writer.writerow([item.name, item.quantity, item.unit])

        # Move to the beginning of the StringIO object
        output.seek(0)

        # Create a Response object with the CSV data
        return Response(
            output.getvalue(),
            mimetype='text/csv',
            headers={
                'Content-Disposition': 'attachment; filename=inventory.csv'
            }
        )
    # Route to download the current inventory as CSV
    @app.route('/invoice-maker', methods=['GET'])
    @login_required
    def invoice_maker():
        return render_template('invoicemaker.html')
    